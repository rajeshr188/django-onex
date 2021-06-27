from django.views.generic import DetailView, ListView, UpdateView, CreateView,DeleteView
from django.views.generic.dates import YearArchiveView,MonthArchiveView,WeekArchiveView
from django.urls import reverse,reverse_lazy
from django.http import HttpResponseRedirect
import re
from django.utils import timezone
from django.shortcuts import render,redirect,get_object_or_404
from django.db.models import Count,Sum,Q,Prefetch
from django.db.models.functions import Cast,Coalesce
from django.db.models.fields import DateField

from .models import License, Series, Loan, Release, Adjustment, Month, Year
from .forms import (LicenseForm, LoanForm, ReleaseForm,Release_formset
                    ,Loan_formset,AdjustmentForm,LoanRenewForm,BulkReleaseForm)
from .tables import LoanTable,ReleaseTable
from .filters import LoanFilter,ReleaseFilter,AdjustmentFilter
from contact.models import Customer

from django_tables2.views import SingleTableMixin
from django_tables2.export.views import ExportMixin
from django_filters.views import FilterView

from num2words import num2words
import math
import datetime
from dateutil.relativedelta import relativedelta
from django.views.decorators.csrf import csrf_exempt

from utils.render import Render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook

@login_required
def print_loanpledge(request,pk):
    loan=Loan.objects.get(id=pk)
    params={'loan':loan}
    return Render.render('girvi/loan_pdf.html',params)

@login_required
def print_release(request,pk):
    release = Release.objects.get(id=pk)
    params = {'release':release}
    return Render.render('girvi/release_pdf.html',params)

@login_required
def print_notice(request,pk):
    pass

@login_required
def notice(request):
    qyr = request.GET.get('qyr',0)
    # qcust = request.GET.GET('qcust',None)
    a_yr_ago = timezone.now() - relativedelta(years=int(qyr))

    cust = Customer.objects.filter(type="Re",loan__created__lt=a_yr_ago,
                                        loan__release__isnull=True).\
                                        prefetch_related(Prefetch('loan_set',
                                                queryset = Loan.objects.filter(
                                                release__isnull=True,
                                                created__lt=a_yr_ago
                                                ))).annotate(
                                                t=Sum('loan__loanamount')
                                                )
    data = {}
    data['totalamount']=cust.aggregate(loancount = Count('loan'),t=Sum('loan__loanamount'))
    data['cust']=cust

    return render(request,'girvi/notice.html',context={'data':data})

@login_required
@csrf_exempt
def multirelease(request,id=None):
    if request.method == 'POST': #<- Checking for method type
        id_list = request.POST.getlist('cbox')
        action = request.POST.get('action')
        print("action" + str(request.POST.get('action')))
        if action == 'delete':
            # delete all selected rows
            print('in delete action')
            Loan.objects.filter(id__in=id_list).delete()
            print('deleted'+ id_list)
        elif action == 'edit':
            print('in edit action'+str(id_list))
            formset = Loan_formset(queryset = Loan.objects.filter(id__in=id_list))
            return render(request, 'girvi/manage_loans.html', {'formset': formset})
        elif action =='release' :
            print('in releaseaction')
            for loan_id in id_list:
                last=Release.objects.all().order_by('id').last()
                if not last:
                    return '1'
                Release.objects.create(releaseid=int(last.id)+1,created=timezone.now(),loan_id=loan_id,interestpaid=0)

    return HttpResponseRedirect(reverse('girvi_loan_list'))

@login_required
def manage_loans(request):

    formset = Loan_formset(queryset=Loan.objects.none())

    if request.method == 'POST':
        formset = formset(request.POST)
        if formset.is_valid():
            # do something with the formset.cleaned_data
            print("saving formset")
            instances=formset.save()
            return redirect(reverse('girvi_loan_list'))
        else :
            print('formset invalid')
            return redirect(reverse('manage_loans',kwargs = {'formset': formset}))
    else:
        formset = Loan_formset(queryset=Loan.objects.none())


    return render(request, 'girvi/manage_loans.html', {'formset': formset})

from dateutil import relativedelta
@login_required
def home(request):
    data = dict()
    today = datetime.date.today()

    loan=dict()
    loans = Loan.objects
    released =Loan.released
    unreleased = Loan.unreleased

    customer=dict()
    c=Customer.objects
    customer['count']=c.count()
    customer['withoutloans']=c.filter(loan__isnull=True).count()
    customer['withloans']=customer['count']-customer['withoutloans']
    customer['latest']=','.join(lat.name for lat in c.order_by('-created')[:5])
    customer['maxloans']=c.filter(loan__release__isnull=True).annotate(num_loans=Count('loan'),sum_loans=Sum('loan__loanamount')).values('name','num_loans','sum_loans').order_by('-num_loans')[:10]

    license =dict()
    l=License.objects.all()
    license['count']=l.count()
    # license['licenses']=','.join(lic.name for lic in l.all())
    series = Series.objects.all()
    license['totalloans']=series.annotate(t=Coalesce(Sum('loan__loanamount',filter=Q(loan__release__isnull=True)),0))
    license['totalunreleasedloans']= series.annotate(t=Coalesce(Sum('loan__loanamount',filter=Q(loan__release__isnull=False)),0))
    license['licchart']=list(license['totalloans'])
    license['licunrchart']=list(license['totalunreleasedloans'])

    loan['total_loans'] = loans.count()
    loan['released_loans'] = released.count()
    loan['unreleased_loans'] = loan['total_loans']-loan['released_loans']

    l=unreleased
    loan['count']=l.count()
    loan['uniquecount']=l.annotate(c=Count('customer')).order_by('c')
    loan['latest']=','.join(lat.loanid for lat in l.order_by('-created')[:5])
    loan['amount']=l.aggregate(t=Coalesce(Sum('loanamount'),0))
    loan['amount_words']=num2words(loan['amount']['t'],lang='en_IN')
    loan['gold_amount']=l.filter(itemtype='Gold').aggregate(t=Sum('loanamount'))
    loan['gold_weight']=l.filter(itemtype='Gold').aggregate(t=Sum('itemweight'))
    loan['gavg']=math.ceil(loan['gold_amount']['t']/loan['gold_weight']['t'])
    loan['silver_amount']=l.filter(itemtype='Silver').aggregate(t=Sum('loanamount'))
    loan['silver_weight']=l.filter(itemtype='Silver').aggregate(t=Sum('itemweight'))
    loan['savg']=math.ceil(loan['silver_amount']['t']/loan['silver_weight']['t'])
    loan['interestdue']= l.aggregate(t=Sum('interest'))


    sumbyitem=(l.aggregate(gold=Sum('loanamount',filter=Q(itemtype="Gold")),silver=Sum('loanamount',filter=Q(itemtype="Silver")),bronze=Sum('loanamount',filter=Q(itemtype="Bronze"))))
    fixed = []
    fixed.append(sumbyitem['gold'])
    fixed.append(sumbyitem['silver'])
    fixed.append(sumbyitem['bronze'])
    loan['sumbyitem']=fixed
    datetimel = loans.annotate(year=Year('created')).values('year').annotate(l = Sum('loanamount')).order_by('year').values_list('year','l',named=True)
    datetime2 = unreleased.annotate(year=Year('created')).values('year').annotate(l = Sum('loanamount')).order_by('year').values_list('year','l',named=True)


    thismonth = loans.filter(created__year = today.year,created__month = today.month)\
                    .annotate(date_only=Cast('created', DateField()))\
                    .values('date_only').annotate(t=Sum('loanamount'))\
                    .order_by('date_only').values_list('date_only','t',named=True)

    lastmonth =  loans.filter(created__year = today.year,created__month = today.month -1)\
                    .annotate(date_only=Cast('created', DateField()))\
                    .values('date_only').annotate(t=Sum('loanamount'))\
                    .order_by('date_only').values_list('date_only','t',named=True)

    thisyear = loans.filter(created__year = today.year).annotate(month=Month('created'))\
                .values('month').order_by('month').annotate(t=Sum('loanamount')).values_list('month','t',named='True')

    lastyear = loans.filter(created__year =today.year -1).annotate(month=Month('created'))\
                .values('month').order_by('month').annotate(t=Sum('loanamount')).values_list('month','t',named='True')
    # lry = Loan.released.dates('created','year').values('created__year').annotate(c = Count('id'))
    loan['status'] = [loans.count(),released.count()]
    
    loan['datechart']=datetimel
    loan['datechart1']=datetime2
    loan['thismonth']=thismonth
    loan['lastmonth']=lastmonth
    loan['thisyear']=thisyear
    loan['lastyear']=lastyear
    release=dict()
    r=Release.objects
    release['count']=r.count()

    data['customer']=customer
    data['license']=license
    data['loan']=loan
    data['release']=release
    return render(request,'girvi/home.html',context={'data':data},)

def check_girvi(request):
    data ={}
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']

        wb=load_workbook(myfile,read_only=True)
        sheet = wb.active

        django_set = set()
        physical_set = set()
        for l in Loan.unreleased.filter(itemtype='Gold').values('loanid'):
            django_set.add(l['loanid'])
        for r in sheet.iter_rows(min_row=0):
            physical_set.add(r[0].value)

        data['records'] = len(django_set)
        data['items'] =  len(physical_set)

        data['missing_records'] = list(physical_set - django_set)
        data['missing_items'] = list(django_set - physical_set)

        return render(request, 'girvi/girvi_upload.html',context={'data':data})

    return render(request, 'girvi/girvi_upload.html')

from django.db import IntegrityError
def bulk_release(request):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = BulkReleaseForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            date = form.cleaned_data['date']
            loans = form.cleaned_data['loans']
            success = []
            failed = []
            if not date:
                date = timezone.now().date()

            for loan in loans:
                try:
                    l = Loan.objects.get(loanid=loan.loanid)
                except Loan.DoesNotExist:
                        # raise CommandError(f"Failed to create Release as {loan} does not exist")
                        print(f"Failed to create Release as {loan} does not exist")
                        continue
                try:
                    releaseid = Release.objects.order_by('-id')[0]
                    releaseid = str(int(releaseid.releaseid)+1)
                    r = Release.objects.create(
                                            releaseid=releaseid,
                                            loan=l,
                                            created = date,#datetime.now(timezone.utc),
                                            interestpaid =  l.interestdue(date)
                                            )
                    success.append(l)
                except IntegrityError:
                    # raise CommandError(f"Failed creating Release as {l} is already Released with {l.release}")
                    failed.append(l)
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            return render(request,'girvi/bulk_release_detail.html',{'success':success,'failed':failed})

    # if a GET (or any other method) we'll create a blank form
    else:
        form = BulkReleaseForm()

    return render(request, 'girvi/bulk_release.html', {'form': form})

def bulk_release_detail(request):
    return render(request,'girvi/bulk_release_detail.html')

class LoanYearArchiveView(LoginRequiredMixin,YearArchiveView):
    queryset = Loan.unreleased.all()
    date_field = "created"
    make_object_list = True
    allow_future = True

class LoanMonthArchiveView(LoginRequiredMixin,MonthArchiveView):
    queryset = Loan.unreleased.all()
    date_field = "created"
    allow_future = True

class LoanWeekArchiveView(LoginRequiredMixin,WeekArchiveView):
    queryset = Loan.unreleased.all()
    date_field = "created"
    week_format = "%W"
    allow_future = True

class LicenseListView(LoginRequiredMixin,ListView):
    model = License

class LicenseCreateView(LoginRequiredMixin,CreateView):
    model = License
    form_class = LicenseForm

class LicenseDetailView(LoginRequiredMixin,DetailView):
    model = License

class LicenseUpdateView(LoginRequiredMixin,UpdateView):
    model = License
    form_class = LicenseForm

class LicenseDeleteView(LoginRequiredMixin,DeleteView):
    model=License
    success_url=reverse_lazy('girvi_license_list')

# class SeriesListView(LoginRequiredMixin,ListView):
#     model = Series
#
# class SeriesCreateView(LoginRequiredMixin,CreateView):
#     model = Series
#     form_class = SeriesForm
#
# class SeriesDetailView(LoginRequiredMixin,DetailView):
#     model = Series
#
# class SeriesUpdateView(LoginRequiredMixin,UpdateView):
#     model = Series
#     form_class = SeriesForm

class LicenseSeriesDeleteView(LoginRequiredMixin,DeleteView):
    model=License
    success_url=reverse_lazy('girvi_license_list')

class LoanListView(LoginRequiredMixin,ExportMixin,SingleTableMixin,FilterView):
    table_class=LoanTable
    model = Loan
    template_name='girvi/loan_list.html'
    filterset_class=LoanFilter
    paginate_by=20

def incloanid():
    last=Loan.objects.all().order_by('id').last()
    if not last:
        return '1'
    splitl=re.split('(\d+)',last.loanid)
    billno=splitl[0] + str(int(splitl[1])+1)
    return str(billno)

def increlid():
    last=Release.objects.all().order_by('id').last()
    if not last:
        return '1'
    # splitl=re.split('(\d+)',last.loanid)
    # billno=splitl[0] + str(int(splitl[1])+1)
    return str(int(last.releaseid)+1)

def ld():
    last=Loan.objects.all().order_by('id').last()
    if not last:
        return datetime.date.today()
    return last.created

class LoanCreateView(LoginRequiredMixin,CreateView):
    model = Loan
    form_class = LoanForm

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        try:
            series ={s.id:list(s.loan_set.values_list('lid').latest('lid')) for s in Series.objects.all()}
            context['series']= series
        except Loan.DoesNotExist:
            context['series']=None
        return context

    def get_initial(self):
        if self.kwargs:
            customer=Customer.objects.get(id=self.kwargs['pk'])
            return{
                'customer':customer,
                'created':ld,
            }
        else:
            return{
                # 'loanid':incloanid,
                'created':ld,
            }

def post_loan(request,pk):
    loan = get_object_or_404(Loan, pk=pk)
    if not loan.posted:
        loan.post()
    return redirect(loan)

def unpost_loan(request, pk):
    loan = get_object_or_404(Loan, pk=pk)
    if loan.posted:
        loan.unpost()
    return redirect(loan)


class LoanDetailView(LoginRequiredMixin,DetailView):
    model = Loan

    def get_context_data(self, **kwargs):
        context = super(LoanDetailView, self).get_context_data(**kwargs)
        context['previous'] = self.object.get_previous()
        context['next'] = self.object.get_next()
        return context

class LoanUpdateView(LoginRequiredMixin,UpdateView):
    model = Loan
    form_class = LoanForm

class LoanDeleteView(LoginRequiredMixin,DeleteView):
    model=Loan
    success_url=reverse_lazy('girvi_loan_list')

def loan_renew(request,pk):
    loan = get_object_or_404(Loan,pk = pk)
    # print(loan.interestdue(date=timezone.now))
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = LoanRenewForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            new_loanamount = loan.due() + form.cleaned_data['amount'] + form.cleaned_data['interest']
            newloan = Loan(series = loan.series,customer = loan.customer,lid = Loan.objects.filter(series=loan.series).latest('lid').lid+1,loanamount=new_loanamount,itemweight=loan.itemweight,itemdesc=loan.itemdesc,interestrate=loan.interestrate)
            Release.objects.create(releaseid = Release.objects.latest('id').id+1,loan=loan,interestpaid=form.cleaned_data['interest'])
            newloan.save()
            # redirect to a new URL:
            return redirect(newloan)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = LoanRenewForm()
    return render(request,'girvi/loan_renew.html',{'form':form,'loan':loan})

class AdjustmentListView(LoginRequiredMixin,ExportMixin,SingleTableMixin,FilterView):
    # table_class=AdjustmentTable
    model = Adjustment
    filterset_class=AdjustmentFilter
    paginate_by=50

class AdjustmentCreateView(LoginRequiredMixin,CreateView):
    model = Adjustment
    form_class = AdjustmentForm
    def get_initial(self):
        if self.kwargs:
            loan=Loan.objects.get(id=self.kwargs['pk'])
            return {'loan':loan,}

class AdjustmentUpdateView(LoginRequiredMixin,UpdateView):
    model = Adjustment
    form_class = AdjustmentForm

class AdjustmentDeleteView(LoginRequiredMixin,DeleteView):
    model=Adjustment
    success_url=reverse_lazy('girvi_adjustments_list')

class ReleaseListView(LoginRequiredMixin,ExportMixin,SingleTableMixin,FilterView):
    table_class=ReleaseTable
    model = Release
    template_name='girvi/release_list.html'
    filterset_class=ReleaseFilter
    paginate_by=10

class ReleaseCreateView(LoginRequiredMixin,CreateView):
    model = Release
    form_class = ReleaseForm

    def get_initial(self):
        if self.kwargs:
            loan=Loan.objects.get(id=self.kwargs['pk'])
            return{'releaseid':increlid,'loan':loan,'interestpaid':loan.interestdue,}

class ReleaseDetailView(LoginRequiredMixin,DetailView):
    model = Release

class ReleaseUpdateView(LoginRequiredMixin,UpdateView):
    model = Release
    form_class = ReleaseForm

class ReleaseDeleteView(LoginRequiredMixin,DeleteView):
    model=Release
    success_url=reverse_lazy('girvi_release_list')
    
def post_release(request, pk):
    release = get_object_or_404(Release, pk=pk)
    if not release.posted:
        release.post()
    return redirect(release)

def unpost_release(request, pk):
    release = get_object_or_404(Release, pk=pk)
    if release.posted:
        release.unpost()
    return redirect(release)