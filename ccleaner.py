from itertools import tee, islice, chain
from openpyxl import Workbook
from openpyxl import load_workbook

import re
import time
# copyright rajeshrathodh @03/2019
start_time = time.time()

Account = []
debtors = []
creditors = []
retailers = []
# wb=load_workbook('Group xlsxLedger - from 01-06-2018 to 31-03-2019.xlsx',read_only=True)
wb = load_workbook('jskdata.xlsx')
sheet = wb.active

txn_wb = Workbook(write_only=True)
txn_ws = txn_wb.create_sheet('sheet')
fr = ['Customer', 'created', 'description', 'docno', 'cash debit',
      'cash credit', 'cash balance', 'metal debit', 'metal credit', 'metal balance']
txn_ws.append(fr)

# txn_inv_ws = txn_wb.create_sheet(title='Invoice')
# txn_rcpt_ws = txn_wb.create_sheet(title='Receipt')
# txn_bal_ws = txn_wb.create_sheet(title = 'Balance')

# inv_wb = Workbook(write_only=True)
# inv_ws = inv_wb.create_sheet('sheet')
# fr = ['id','customer', 'created', 'rate','description', 'DocNo',
#       'balancetype', 'paymenttype', 'balance']
# inv_ws.append(fr)

# rcpt_wb = Workbook(write_only=True)
# rcpt_ws = rcpt_wb.create_sheet('sheet')
# fr = ['id','customer', 'created', 'description', 'DocNo',
#       'type', 'total','status']
# rcpt_ws.append(fr)

bal_wb = Workbook(write_only=True)
bal_ws = bal_wb.create_sheet('sheet')
fr = ['Customer','type', 'Amount Bal', 'Gold Bal']
bal_ws.append(fr)

for row in sheet.iter_rows(min_row=2):
    if row[0].value is not None:
        if 'Account Name:' in row[0].value:
            Account.append(row[0].coordinate)
        if 'Sundry Debtors' in row[0].value:
            debtors.append(row[0].coordinate)
        elif 'Sundry Creditors' in row[0].value:
            creditors.append(row[0].coordinate)
        elif 'Retail Customer' in row[0].value:
            retailers.append(row[0].coordinate)

def previous_and_next(some_iterable):
    prevs, items, nexts = tee(some_iterable, 3)
    prevs = chain([None], prevs)
    nexts = chain(islice(nexts, 1, None), [None])
    return zip(prevs, items, nexts)

for previous, item, nxt in previous_and_next(Account):
    # print("Item is now", item, sheet[str(item)].coordinate, sheet[str(item)].value, "next is", nxt, "previous is", previous)
    if nxt is None:
        nxt = 'A'+str(sheet.max_row-1)

    name = re.search(r'\)\s([^)]+)', sheet[str(item)].value).group(1).strip()
    atype = re.search(r'\(([^)]+)', sheet[str(item)].value).group(1).strip()
    tt = ''
    if atype == 'Sundry Creditors':
        tt = 'Su'
    elif atype == 'Sundry Debtors':
        tt = 'Wh'
    else:
        tt = 'Re'
    # print(name,tt)
    arange = str(sheet[str(item)].coordinate +
                 ":"+"A"+sheet[str(nxt)].coordinate)

    row = sheet[str(item)].row
    col = sheet[str(item)].column
    nxt_coordinate = sheet[nxt].coordinate
    nxt_row = sheet[str(nxt)].row
    nxt_col = sheet[str(nxt)].column

    balance = [name, tt, sheet.cell(
        row=nxt_row-2, column=7).value, sheet.cell(row=nxt_row-2, column=10).value] 
    bal_ws.append(balance)

    # for cell in sheet[f"B{row+1}:AA{nxt_row-4}"]:
    #     r = cell[0].row
    #     tdict = {}

    #     tdict['date'] = sheet['B'+str(r)].value
    #     tdict['description'] = sheet['C'+str(r)].value
    #     tdict['docno'] = sheet['D'+str(r)].value
    #     tdict['debitamount'] = sheet['E'+str(r)].value
    #     tdict['creditamount'] = sheet['F'+str(r)].value
    #     tdict['balanceamount'] = sheet['G'+str(r)].value
    #     tdict['debitmetal'] = sheet['H'+str(r)].value
    #     tdict['creditmetal'] = sheet['I'+str(r)].value
    #     tdict['balancemetal'] = sheet['J'+str(r)].value
    #     txn = [name,tdict['date'], tdict['description'], tdict['docno'],
    #             tdict['debitamount'],tdict['creditamount'],tdict['balanceamount'],
    #             tdict['debitmetal'],tdict['creditmetal'],tdict['balancemetal']]
    #     txn_ws.append(txn)
    #     if tdict['debitamount'] is not None:
    #         inva = ['',name,tdict['date'],'',tdict['description'],tdict['docno'],'Cash','Credit',tdict['debitamount']]
    #         inv_ws.append(inva)
    #     if tdict['creditamount'] is not None:
    #         rcpa = ['',name, tdict['date'], tdict['description'],
    #                tdict['docno'], 'Cash', tdict['creditamount']]
    #         rcpt_ws.append(rcpa)
    #     if tdict['debitmetal'] is not None:
    #         invm = ['',name, tdict['date'], '',tdict['description'],
    #                tdict['docno'], 'Metal', 'Credit', tdict['debitmetal']]
    #         inv_ws.append(invm)
    #     if tdict['creditmetal'] is not None:
    #         rcpm = ['',name, tdict['date'], tdict['description'],
    #                tdict['docno'], 'Metal', tdict['creditmetal']]
    #         rcpt_ws.append(rcpm)

# txn_inv_ws = inv_wb.copy_worksheet(inv_ws)
# txn_rcpt_ws = rcpt_wb.copy_worksheet(rcpt_ws)
# txn_bal_ws = bal_wb.copy_worksheet(bal_ws)

# txn_wb.save(filename='clean_data_txn.xlsx')
# inv_wb.save(filename='clean_data_inv.xlsx')
# rcpt_wb.save(filename='clean_data_rcpt.xlsx')
bal_wb.save(filename='clean_data_bal.xlsx')

print("--- %s seconds ---" % (time.time() - start_time))
