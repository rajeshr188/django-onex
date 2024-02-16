import base64
import datetime
import math
import textwrap

from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.db.models import (Count, F, Max, OuterRef, Prefetch, Q, Subquery,
                              Sum)
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.response import TemplateResponse
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import (require_http_methods,  # new
                                          require_POST)
from django.views.generic.dates import (DayArchiveView, MonthArchiveView,
                                        TodayArchiveView, WeekArchiveView,
                                        YearArchiveView)
from django_tables2.config import RequestConfig
from django_tables2.export.export import TableExport
from dynamic_preferences.registries import global_preferences_registry
from num2words import num2words
from openpyxl import load_workbook
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

from contact.models import Customer
from notify.models import NoticeGroup, Notification
from rates.models import Rate
from utils.htmx_utils import for_htmx
from utils.loan_pdf import get_loan_pdf, get_notice_pdf

from ..filters import LoanFilter
from ..forms import LoanForm, LoanItemForm, LoanRenewForm
from ..models import *
from ..tables import LoanTable

# We instantiate a manager for our global preferences
global_preferences = global_preferences_registry.manager()


def create_loan_notification(request, pk=None):
    # get loan instance
    loan = get_object_or_404(Loan, pk=pk)
    # create a noticegroup
    import random
    import string

    # Generate a random string of 3 letters
    random_string = "".join(random.choice(string.ascii_letters) for _ in range(3))
    ng = NoticeGroup.objects.create(
        name=f"{loan.loan_id}-{random_string}-{datetime.now().date()}"
    )
    notification = Notification.objects.create(
        group=ng,
        customer=loan.customer,
    )
    # add the loan to the notification
    notification.loans.add(loan)
    notification.save()
    return redirect(notification.get_absolute_url())


class LoanYearArchiveView(LoginRequiredMixin, YearArchiveView):
    queryset = Loan.objects.unreleased()
    date_field = "loan_date"
    make_object_list = True


class LoanMonthArchiveView(LoginRequiredMixin, MonthArchiveView):
    queryset = Loan.objects.unreleased()
    date_field = "loan_date"
    make_object_list = True

    def get_context_data(self, *args, **kwargs):
        data = super().get_context_data(**kwargs)
        data["count"] = len(data)
        return data


class LoanWeekArchiveView(LoginRequiredMixin, WeekArchiveView):
    queryset = Loan.objects.unreleased()
    date_field = "loan_date"
    week_format = "%W"


class LoanDayArchiveView(LoginRequiredMixin, DayArchiveView):
    queryset = Loan.objects.all()
    date_field = "loan_date"
    allow_empty = True


class LoanTodayArchiveView(TodayArchiveView):
    queryset = Loan.objects.all()
    date_field = "loan_date"
    allow_empty = True
    # template_name = "girvi/loan/loan_archive_day.html"


def ld():
    default_date = global_preferences["Loan__Default_Date"]
    if default_date == "N":
        return datetime.now()
    else:
        last = Loan.objects.order_by("id").last()
        if not last:
            return datetime.now()
        return last.loan_date


def get_interestrate(request):
    metal = request.GET["itemtype"]
    interest = 0
    if metal == "Gold":
        interest = global_preferences["Interest_Rate__gold"]
    elif metal == "Silver":
        interest = global_preferences["Interest_Rate__silver"]
    else:
        interest = global_preferences["Interest_Rate__other"]

    form = LoanItemForm(initial={"interestrate": interest})
    context = {
        "field": form["interestrate"],
    }
    return render(request, "girvi/partials/field.html", context)


@login_required
@for_htmx(use_block="content")
def loan_list(request):
    stats = {}
    filter = LoanFilter(
        request.GET,
        queryset=Loan.objects.order_by("-id")
        .with_details()
        # .select_related("customer", "release", "created_by")
        .prefetch_related("notifications", "loanitems"),
    )
    table = LoanTable(filter.qs)

    RequestConfig(request, paginate={"per_page": 10}).configure(table)
    export_format = request.GET.get("_export", None)
    if TableExport.is_valid_format(export_format):
        # TODO speed up the table export using celery

        exporter = TableExport(
            export_format,
            table,
            exclude_columns=(
                "selection",
                "notified",
                "months_since_created",
                "current_value",
                "total_due",
                "total_interest",
            ),
            dataset_kwargs={"title": "loans"},
        )
        return exporter.response(f"table.{export_format}")
    context = {"filter": filter, "table": table}
    return TemplateResponse(request, "girvi/loan/loan_list.html", context)


@login_required
@for_htmx(use_block="content")
def create_loan(request, pk=None):
    if request.method == "POST":
        form = LoanForm(request.POST, request.FILES)
        if form.is_valid():
            l = form.save(commit=False)
            l.created_by = request.user

            image_data = request.POST.get("image_data")
            if image_data:
                image_file = ContentFile(
                    base64.b64decode(image_data.split(",")[1]),
                    name=f"{l.loan_id}__{l.customer.name}_{l.id}.jpg",
                )
                l.pic = image_file

            l.save()
            messages.success(request, f"Created Loan : {l.loan_id}")

            response = TemplateResponse(
                request,
                "girvi/loan/loan_detail.html",
                {"loan": l, "object": l},
            )
            response["Hx-Push-Url"] = reverse(
                "girvi:girvi_loan_detail", kwargs={"pk": l.id}
            )
            return response

        messages.warning(request, "Please correct the error below.")
        response = TemplateResponse(
            request, "girvi/loan/loan_form.html", {"form": form}
        )
        return response

    try:
        series = Loan.objects.latest().series
        lid = series.loan_set.last().lid + 1
    except Loan.DoesNotExist:
        series = Series.objects.first()
        lid = 1

    initial = {
        "loan_date": ld(),
        "series": series,
        "lid": lid,
    }
    if pk:
        initial["customer"] = get_object_or_404(Customer, pk=pk)

    form = LoanForm(initial=initial)
    return TemplateResponse(request, "girvi/loan/loan_form.html", {"form": form})


@login_required
@for_htmx(use_block="content")
def loan_update(request, id=None):
    obj = get_object_or_404(Loan, id=id)
    form = LoanForm(request.POST or None, instance=obj)

    if form.is_valid():
        l = form.save(commit=False)
        l.created_by = request.user
        image_data = request.POST.get("image_data")
        if image_data:
            image_file = ContentFile(
                base64.b64decode(image_data.split(",")[1]),
                name=f"{l.loan_id}__{l.customer.name}_{l.id}.jpg",
            )
            l.pic = image_file

        l.save()
        messages.success(request, messages.SUCCESS, f"updated Loan {l.loan_id}")

        response = TemplateResponse(
            request, "girvi/loan/loan_detail.html", {"loan": obj, "object": obj}
        )
        response["Hx-Push-Url"] = reverse(
            "girvi:girvi_loan_detail", kwargs={"pk": obj.id}
        )
        return response

    return TemplateResponse(
        request, "girvi/loan/loan_form.html", {"form": form, "loan": obj, "object": obj}
    )


@require_http_methods(["DELETE"])
@login_required
def loan_delete(request, pk=None):
    obj = get_object_or_404(Loan, id=pk)
    obj.delete()
    messages.error(request, f" Loan {obj} Deleted")
    return HttpResponse(
        status=204, headers={"Hx-Redirect": reverse("girvi:girvi_loan_list")}
    )


@login_required
def next_loanid(request):
    try:
        series = request.GET.get("series")
        s = get_object_or_404(Series, pk=series)

        last_loan = s.loan_set.last()
        if last_loan:
            lid = last_loan.lid + 1
        else:
            lid = 1

        form = LoanForm(initial={"lid": lid})
        context = {
            "field": form["lid"],
        }
        return render(request, "girvi/partials/field.html", context)
    except (Series.DoesNotExist, Exception) as e:
        # Handle exceptions here, you can log the error or return an error response
        # For simplicity, here we are returning a basic error message
        return render(
            request,
            "error.html",
            {"error_message": "An error occurred in next_loanid."},
        )


@login_required
@for_htmx(use_block_from_params=True)
def loan_detail(request, pk):
    loan = get_object_or_404(
        Loan.objects.select_related(
            "customer", "series", "release", "created_by"
        ).prefetch_related("loan_payments", "loanitems"),
        pk=pk,
    )
    result = []
    for item in loan.get_weight:
        item_type = item["itemtype"]
        total_weight_purity = round(item["total_weight"])
        result.append(f"{item_type}:{total_weight_purity}")

    # Join the results into a single string
    weight = " ".join(result)
    result = {}
    for item in loan.get_pure:
        item_type = item["itemtype"]
        total_weight_purity = round(item["pure_weight"], 3)
        result[item["itemtype"]] = total_weight_purity

    # Subquery to get the latest timestamp for each metal type
    latest_timestamps = (
        Rate.objects.filter(metal=OuterRef("metal"))
        .order_by("-timestamp")
        .values("timestamp")[:1]
    )

    # Query to get the latest rate for each metal type
    latest_rates = Rate.objects.filter(
        timestamp__in=Subquery(latest_timestamps)
    ).values("metal", "buying_rate")

    # Create a dictionary to store the latest rates for each metal type
    rate_dict = {rate["metal"]: rate["buying_rate"] for rate in latest_rates}

    # Calculate the values using the latest rates from the database
    result_dict = {
        itemtype: rate_dict.get(itemtype, None) * weight
        for itemtype, weight in result.items()
    }
    # Join the results into a single string

    value = round(sum(result_dict.values()))
    context = {
        "object": loan,
        "sunken": loan.total() < value,
        "items": loan.loanitems.all(),
        "statements": loan.statementitem_set.all(),
        "loan": loan,
        "weight": weight,  # loan.get_weight,
        "pure": result,
        "value": value,
        "worth": value - loan.due(),
        # "journal_entries": loan.journal_entries.all(),
        "new_item_url": reverse(
            "girvi:girvi_loanitem_create", kwargs={"parent_id": loan.id}
        ),
    }

    return TemplateResponse(request, "girvi/loan/loan_detail.html", context)


@login_required
@for_htmx(use_block="content")
def loan_renew(request, pk):
    loan = get_object_or_404(Loan, pk=pk)

    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = LoanRenewForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            new_loanamount = (
                loan.due() + form.cleaned_data["amount"] + form.cleaned_data["interest"]
            )
            newloan = Loan(
                series=loan.series,
                customer=loan.customer,
                lid=Loan.objects.filter(series=loan.series).latest("lid").lid + 1,
                loanamount=new_loanamount,
            )
            newloan.save()
            # copy and create loan.loanitems to newloan.items
            for item in loan.loanitems.all():
                newloan.loanitems.add(item)
            # create a new release object
            Release.objects.create(
                releaseid=Release.objects.latest("id").id + 1,
                loan=loan,
                interestpaid=form.cleaned_data["interest"],
            )
            newloan.save()
            # redirect to a new URL:
            messages.success(request, f"Renewed Loan : {newloan.loanid}")
            return redirect(newloan)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = LoanRenewForm()
    response = TemplateResponse(
        request, "girvi/loan/loan_renew.html", {"form": form, "loan": loan}
    )
    response["Hx-Push-Url"] = reverse("girvi:girvi_loan_renew", kwargs={"pk": loan.id})
    return response


@login_required
def notice(request):
    qyr = request.GET.get("qyr", 0)

    a_yr_ago = timezone.now() - relativedelta(years=int(qyr))

    # get all loans with selected ids
    selected_loans = (
        Loan.objects.unreleased()
        .filter(loan_date__lt=a_yr_ago)
        .order_by("customer")
        .select_related("customer")
    )

    # get a list of unique customers for the selected loans
    # customers = selected_loans.values('customer').distinct().count()
    customers = (
        Customer.objects.filter(loan__in=selected_loans)
        .distinct()
        .prefetch_related("loan_set", "address", "contactno")
    )

    data = {}
    data["loans"] = selected_loans
    data["loancount"] = selected_loans.count()
    data["total"] = selected_loans.total_loanamount()
    data["interest"] = selected_loans.with_total_interest()
    data["cust"] = customers

    return render(request, "girvi/loan/notice.html", context={"data": data})


@require_http_methods("POST")
@login_required
def deleteLoan(request):
    id_list = request.POST.getlist("selection")
    loans = Loan.objects.filter(id__in=id_list)
    for i in loans:
        i.delete()
    messages.error(request, f"Deleted {len(id_list)} loans")
    return loan_list(request)


@login_required
def notify_print(request):
    # check if user wanted all rows to be selected
    all = request.POST.get("selectall")
    selected_loans = None

    if all == "selected":
        print("all selected")
        # get query parameters if all row selected and retrive queryset
        print(request.GET)
        filter = LoanFilter(
            request.GET,
            queryset=Loan.objects.unreleased()
            .select_related("customer", "release")
            .prefetch_related("notifications", "loanitems"),
        )

        selected_loans = filter.qs.order_by("customer")
        print(f"selected loans: {selected_loans.count()}")
    else:
        print("partially selected")
        # get the selected loan ids from the request
        selection = request.POST.getlist("selection")

        selected_loans = (
            Loan.objects.unreleased().filter(id__in=selection).order_by("customer")
        )

    if selected_loans:
        # Create a new NoticeGroup
        ng = NoticeGroup.objects.create(name=datetime.datetime.now())

        # Get a queryset of customers with selected loans
        customers = Customer.objects.filter(loan__in=selected_loans).distinct()

        # Create a list of Notification objects to create
        notifications_to_create = []
        for customer in customers:
            notifications_to_create.append(
                Notification(
                    group=ng,
                    customer=customer,
                )
            )
        # Use bulk_create to create the notifications
        try:
            notifications = Notification.objects.bulk_create(notifications_to_create)
        except IntegrityError:
            print("Error adding notifications.")

        # Add loans to the notifications
        for notification in notifications:
            loans = selected_loans.filter(customer=notification.customer)
            notification.loans.set(loans)
            notification.save()
        return redirect(ng.get_absolute_url())

    return HttpResponse(status=200, content="No unreleased loans selected.")


@login_required
def print_loan(request, pk=None):
    loan = get_object_or_404(Loan, pk=pk)
    pdf = get_loan_pdf(loan=loan)
    # Create a response object
    response = HttpResponse(pdf, content_type="application/pdf")
    # response["Content-Disposition"] = 'attachment; filename="pledge.pdf"'
    response["Content-Disposition"] = f"inline; filename='{loan.lid}.pdf'"
    return response


@login_required
def generate_original(request, pk=None):
    loan = get_object_or_404(Loan, pk=pk)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename='{loan.lid}.pdf'"

    page_width = 14.6 * cm
    page_height = 21 * cm
    c = canvas.Canvas(response, pagesize=(page_width, page_height))

    # Grid spacing
    grid_spacing = 1 * cm  # Adjust this value based on your preference
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(12 * cm, 18.2 * cm, f"{loan.lid}")
    c.drawString(11.5 * cm, 17.2 * cm, f"{loan.loan_date.strftime('%d-%m-%Y')}")

    # Wrap the text if its length is greater than 35 characters
    customer = f"{loan.customer.name} {loan.customer.get_relatedas_display()} {loan.customer.relatedto}"
    lines = textwrap.wrap(customer, width=35)
    # Draw the wrapped text on the canvas
    y = 15.5 * cm
    for line in lines:
        c.drawString(2 * cm, y, line)
        y -= 0.5 * cm
    c.setFont("Helvetica", 12)
    address = f"{loan.customer.address.first()}"
    lines = textwrap.wrap(address, width=35)
    # Draw the wrapped text on the canvas
    y = 15 * cm
    for line in lines:
        c.drawString(2 * cm, y, line)
        y -= 0.5 * cm
    c.setFont("Helvetica", 12)
    c.drawString(2 * cm, 14 * cm, f"Ph: {loan.customer.contactno.first()}")

    result = []
    for item in loan.get_weight:
        item_type = item["itemtype"]
        total_weight_purity = item["total_weight"]
        result.append(f"{item_type}:{total_weight_purity}")

    # Join the results into a single string
    weight = " ".join(result)

    c.setFont("Helvetica", 8)
    c.drawString(10 * cm, 11.8 * cm, f"{weight}gms ")
    result = []
    for item in loan.get_pure:
        item_type = item["itemtype"]
        total_weight_purity = round(item["pure_weight"])
        result.append(f"{item_type}:{total_weight_purity}")
    # Join the results into a single string
    pure = " ".join(result)
    c.drawString(10 * cm, 9.8 * cm, f"{pure} gms ")
    c.drawString(10 * cm, 8 * cm, f"{loan.current_value()}")

    c.setFont("Helvetica-Bold", 14)
    # c.drawString(7 * cm, 12.5 * cm, f"{loan.itemtype}")
    c.drawString(8 * cm, 6.5 * cm, f"{loan.loan_amount}")
    # Wrap the text if its length is greater than 35 characters
    lines = textwrap.wrap(loan.item_desc, width=30)
    # Draw the wrapped text on the canvas
    y = 11 * cm
    for line in lines:
        c.drawString(2.5 * cm, y, line)
        y -= 0.5 * cm

    c.setFont("Helvetica", 12)
    c.drawString(
        2 * cm, 6 * cm, f"{num2words(loan.loan_amount, lang='en_IN')} rupees only"
    )
    c.showPage()
    c.setFont("Helvetica", 12)
    # Grid spacing
    grid_spacing = 1 * cm  # Adjust this value based on your preference

    c.drawString(3 * cm, 17 * cm, f"{loan.lid}")
    c.drawString(11 * cm, 17 * cm, f"{loan.loan_date.strftime('%d-%m-%Y')}")
    c.drawString(5 * cm, 16.5 * cm, f"{loan.customer.name}")
    c.drawString(
        5 * cm,
        16 * cm,
        f"{loan.customer.get_relatedas_display()} {loan.customer.relatedto}",
    )
    address = textwrap.wrap(f"{loan.customer.address.first()}")
    # Draw the wrapped text on the canvas
    y = 15 * cm
    for line in address:
        c.drawString(5 * cm, y, line)
        y -= 0.5 * cm

    c.drawString(5 * cm, 14 * cm, f"{loan.customer.contactno.first()}")
    c.drawString(5 * cm, 13.5 * cm, f"{loan.loan_amount}")

    lines = textwrap.wrap(
        f"{num2words(loan.loan_amount, lang='en_IN')} rupees only", width=45
    )

    # Wrap the text if its length is greater than 35 characters
    lines = textwrap.wrap(loan.item_desc, width=35)
    # Draw the wrapped text on the canvas
    y = 10 * cm
    for line in lines:
        c.drawString(3 * cm, y, line)
        y -= 0.5 * cm

    c.setFont("Helvetica", 8)
    result = []
    for item in loan.get_weight:
        item_type = item["itemtype"]
        total_weight_purity = round(item["total_weight"])
        result.append(f"{item_type}:{total_weight_purity}")

    # Join the results into a single string
    weight = " ".join(result)
    c.drawString(3 * cm, 7 * cm, f"{weight}gms")
    result = []
    for item in loan.get_pure:
        item_type = item["itemtype"]
        total_weight_purity = round(item["pure_weight"])
        result.append(f"{item_type}:{total_weight_purity}")

    # Join the results into a single string
    pure = " ".join(result)
    c.drawString(7 * cm, 7 * cm, f"{pure}")
    c.setFont("Helvetica", 12)
    c.drawString(12 * cm, 7 * cm, f"{loan.current_value()}")
    c.setFont("Helvetica-Bold", 12)
    c.drawString(1 * cm, 3.5 * cm, f"{loan.lid}")
    c.drawString(1 * cm, 3 * cm, f"{loan.loan_date.strftime('%d/%m/%y')}")
    c.drawString(1 * cm, 2.5 * cm, f"{loan.loan_amount}   {weight}")
    c.drawString(1 * cm, 2 * cm, f"{loan.customer.name}")
    # Wrap the text if its length is greater than 35 characters
    lines = textwrap.wrap(f"{loan.item_desc}", width=35)
    # Draw the wrapped text on the canvas
    y = 1.5 * cm
    for line in lines:
        c.drawString(1 * cm, y, line)
        y -= 0.5 * cm

    c.save()
    return response


# @user_passes_test(lambda user: user.is_superuser)

from django.db import transaction
from django.http import HttpResponse, JsonResponse


def statement_create(request):
    if request.method == "POST":
        try:
            myfile = request.FILES["myfile"]
            wb = load_workbook(myfile, read_only=True)
            sheet = wb.active
        except KeyError:
            return JsonResponse(
                {"error": "file field is missing in the request."}, status=400
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

        with transaction.atomic():
            statement = Statement.objects.create(created_by=request.user)
            statement_items = []

            for row in sheet.iter_rows(min_row=0):
                loan_id = row[0].value
                loan = Loan.objects.filter(loanid=loan_id).first()

                if loan:
                    statement_items.append(
                        StatementItem(statement=statement, loan=loan)
                    )

            StatementItem.objects.bulk_create(statement_items)

        return JsonResponse(
            {
                "message": f"Statement {statement} with {len(statement_items)} items created."
            },
            status=201,
        )

    return JsonResponse(
        {"error": "This endpoint only accepts POST requests."}, status=405
    )


@login_required
def check_girvi(request, pk=None):
    if pk:
        statement = get_object_or_404(Statement, pk=pk)
    else:
        statement = Statement.objects.last()

    unreleased = list(Loan.objects.unreleased().values_list("loanid", flat=True))
    unreleased_set = set(unreleased)
    physical = list(
        StatementItem.objects.filter(statement=statement).values_list(
            "loan__loanid", flat=True
        )
    )
    statement_set = set(physical)
    data = {}
    data["records"] = len(unreleased_set)
    data["items"] = len(statement_set)

    data["missing_records"] = list(statement_set - unreleased_set)
    data["missing_items"] = list(unreleased_set - statement_set)
    missing_records = Loan.objects.filter(loanid__in=data["missing_records"])
    missing_items = Loan.objects.filter(loanid__in=data["missing_items"])
    return render(
        request,
        "girvi/loan/girvi_upload.html",
        context={
            "data": data,
            "statement": statement,
            "missing_items": missing_items,
            "missing_records": missing_records,
        },
    )


@login_required
def loan_item_update_hx_view(request, parent_id=None, id=None):
    if not request.htmx:
        raise Http404("This view is meant for Htmx requests only.")

    try:
        parent_obj = Loan.objects.get(id=parent_id)
    except Loan.DoesNotExist:
        return HttpResponse("Not found.", status=404)

    instance = None
    if id is not None:
        try:
            instance = LoanItem.objects.get(loan=parent_obj, id=id)
        except LoanItem.DoesNotExist:
            instance = None

    if request.method == "POST":
        form = LoanItemForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            new_obj = form.save(commit=False)
            if instance is None:
                new_obj.loan = parent_obj
            image_data = request.POST.get("image_data")

            if image_data:
                image_file = ContentFile(
                    base64.b64decode(image_data.split(",")[1]),
                    name=f"{new_obj.loan.loan_id}_{new_obj.id}.jpg",
                )

                new_obj.pic = image_file
            new_obj.save()
            messages.success(request, f"Created Item : {new_obj.id}")
            context = {"object": new_obj}

            if request.htmx:
                return HttpResponse(status=204, headers={"HX-Trigger": "loanChanged"})
            return render(request, "girvi/partials/item-inline.html", context)
        # else:
        #     context = {"url": url, "form": form, "object": instance}
        #     return render(request, "girvi/partials/item-form.html", context)

    else:
        form = LoanItemForm(instance=instance)

    url = reverse("girvi:girvi_loanitem_create", kwargs={"parent_id": parent_obj.id})
    if instance:
        url = instance.get_hx_edit_url()

    context = {"url": url, "form": form, "object": instance}
    return render(request, "girvi/partials/item-form.html", context)


# @login_required
# def get_loan_items(request, loan):
#     l = get_object_or_404(Loan, pk=loan)
#     items = l.loanitem_set.all()
#     return render(request, "", context={"items": items})


@login_required
def loanitem_delete(request, parent_id, id):
    item = get_object_or_404(LoanItem, id=id, loan_id=parent_id)
    loan = item.loan
    item.delete()
    messages.error(request, f"Item {item} Deleted")
    loan.save()
    return HttpResponse(status=204, headers={"HX-Trigger": "loanChanged"})


@login_required
def loanitem_detail(request, pk):
    item = get_object_or_404(LoanItem, pk=pk)
    return render(request, "girvi/partials/item-inline.html", {"object": item})
