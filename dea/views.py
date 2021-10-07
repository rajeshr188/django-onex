from dea.tables import AccountbalanceTable
from typing import List
from django_tables2.views import SingleTableMixin
from openpyxl import load_workbook
from django.db import transaction
from django.db.models.query_utils import subclasses
from django.shortcuts import get_object_or_404, render, redirect
from django.urls.base import reverse_lazy
from django.views.generic import CreateView, ListView, DeleteView, DetailView
from django.shortcuts import render
from django_filters.views import FilterView
from django_tables2.export.views import ExportMixin
from django.contrib.auth.mixins import LoginRequiredMixin


from.filters import AccountFilter,LedgerFilter
# Create your views here.
from .models import Account, AccountStatement, Accountbalance,Journal, Ledger, LedgerStatement, LedgerTransaction,Ledgerbalance
from .forms import AccountForm, AccountStatementForm, LedgerForm, LedgerStatementForm
from dea.utils.currency import Balance
import logging

logger = logging.getLogger(__name__)
def home(request):
    lb = Ledgerbalance.objects.all()
    
    context={}
    balancesheet = {}
    balancesheet['assets'] = lb.filter(AccountType = 'Asset')
    
    ta = Balance()
    tl = Balance()
    ti = Balance()
    te = Balance()
    for i in lb:
        if i.AccountType == 'Asset':
            ta = ta+ abs(i.get_currbal())
        elif i.AccountType == 'Liability':
            tl = tl+ abs(i.get_currbal())
        elif i.AccountType == 'Income':
            ti = ti + abs(i.get_currbal())
        elif i.AccountType == 'Expense':
            te = te + abs(i.get_currbal())
    pnloss = {}
    pnloss['income'] = lb.filter(AccountType = 'Income')
    pnloss['expense'] = lb.filter(AccountType = 'Expense')
    context['ta'] = ta
    context['tl'] = tl
    context['ti'] = ti
    context['te'] = te
    balancesheet['liabilities'] = lb.filter(AccountType = 'Liability')
    context['pnloss'] = pnloss
        
    context['balancesheet'] = balancesheet
    context['ledger']=lb   
                      
    # context['accounts'] = AccountStatement.objects.filter(
    #                 pk__in = AccountStatement.objects.order_by().values('AccountNo').annotate(max_id = Max('id')).values('max_id')).select_related("AccountNo","AccountNo__contact")
    # a = Account.objects.filter(contact__type='Su')
    context['accounts'] = []
    
    context['journal'] = []
    # jrnls = []
    # for i in journal:
    #     # txns = lt.filter(journal = i,)
    #     jrnls.append({
    #                     'object_id':i.object_id,
    #                     'id':i.id,'created':i.created,'desc':i.desc,
    #                     'app_label':i.content_type.app_label,'model':i.content_type.model,
    #                     'content_object':i.content_object,'url_string':i.get_url_string(),
    #                     'content_type':i.content_type,
    #                     # 'txns':txns
    #                     })
    
    return render(request, 'dea/home.html', {'data': context})

def generalledger(request):
    context = {}
    context['lt'] = LedgerTransaction.objects.all().order_by('-created').select_related('journal','ledgerno','ledgerno_dr')
    return render(request,'dea/gl.html',{'data':context})

def daybook(request):
    try:
        latest_stmt = LedgerStatement.objects.latest()
    except:
        print("no ledger statements")

def set_ledger_ob(request, pk):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = LedgerStatementForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            form.save()
            return redirect('/')

    # if a GET (or any other method) we'll create a blank form
    else:
        form = LedgerStatementForm()

    return render(request, 'dea/set_ledger_ob.html', {'form': form})

def set_acc_ob(request, pk):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = AccountStatementForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            form.save()
            return redirect('/')

    # if a GET (or any other method) we'll create a blank form
    else:
        form = AccountStatementForm()

    return render(request, 'dea/set_acc_ob.html', {'form': form})

@transaction.atomic()
def audit_acc(request,pk):
    acc = get_object_or_404(Account,pk=pk)
    acc.audit()
    return redirect(acc)

@transaction.atomic()
def audit_ledger(request):
    ledgers = Ledger.objects.all()
    for l in ledgers:
        l.audit()
    return redirect("/dea")
    
class JournalListView(ListView):
    queryset = Journal.objects.all().select_related('content_type')
    paginate_by = 10

class JournalDetailView(DetailView):
    model = Journal

class JournalDeleteView(DeleteView):
    model = Journal
    success_url = reverse_lazy('dea_journals_list')

class AccountCreateView(CreateView):
    model = Account
    form_class = AccountForm

class AccountListView(LoginRequiredMixin,ExportMixin,SingleTableMixin,FilterView):
    # queryset = Account.objects.all().select_related('entity','AccountType_Ext','contact')
    table_class = AccountbalanceTable
    model = Accountbalance
    filterset_class = AccountFilter
    template_name = 'dea/account_list.html'
    paginate_by = 10

class AccountDetailView(DetailView):
    model = Account

    def get_context_data(self, **kwargs):
        ct = super().get_context_data(**kwargs)
        acc = ct['object']
        if ct['object'].accountstatements.exists():
            acc_stmt = acc.accountstatements.latest()
            ls_created = acc_stmt.created 
            txns = list(acc.txns(since=ls_created))
            
        else:
            txns = list(acc.txns())
            acc_stmt = None

        ct['acc_stmt'] = acc_stmt
        op_bal = Balance() if acc_stmt is None else acc_stmt.get_cb()
        running_bal = op_bal
        txn_list = []
        for i in txns:
            txn_dict = {}
            txn_dict['created'] = i.created
            txn_dict['voucher'] = i.journal.content_object
            txn_dict['voucher_url'] = i.journal.get_url_string()
            txn_dict['description'] = i.XactTypeCode_ext.description
            txn_dict['xactcode'] = i.XactTypeCode.XactTypeCode
            txn_dict['amount'] = i.amount
            if i.XactTypeCode.XactTypeCode == 'Dr':
               
                txn_dict['running_bal'] = running_bal + Balance([i.amount])
            else:
                
                txn_dict['running_bal'] = running_bal - Balance([i.amount])
            txn_list.append(txn_dict)
        ct['txns'] = txn_list
        return ct

class AccountStatementCreateView(CreateView):
    model = Account
    form_class = AccountStatementForm

class AccountStatementListView(ListView):
    model = AccountStatement

class AccountStatementDeleteView(DeleteView):
    model = AccountStatement
    success_url = reverse_lazy('dea_account_list')

class LedgerCreateView(CreateView):
    model = Ledger
    form_class = LedgerForm

class LedgerListView(ListView):
    model = Ledger

class LedgerDetailView(DetailView):
    model = Ledger

    def get_context_data(self, **kwargs) :
        ct = super().get_context_data(**kwargs)
        ls_created = ct['object'].ledgerstatements.latest().created if ct['object'].ledgerstatements.exists() else None
        ct['dtxns']= ct['object'].dtxns(since=ls_created)
        ct['ctxns']= ct['object'].ctxns(since = ls_created)
        return ct

class LedgerStatementCreateView(CreateView):
    model = LedgerStatement
    form_class = LedgerStatementForm

class LedgerStatementListView(ListView):
    model = LedgerStatement

class LedgerTransactionListView(ListView):
    model=LedgerTransaction
    queryset = LedgerTransaction.objects.select_related('ledgerno','ledgerno_dr','journal')
    template_name = 'dea/gl.html'
    paginate_by = 20

from openpyxl import load_workbook
from moneyed import Money
from .utils.currency import Balance
from decimal import Decimal
def import_acc_opbal(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        wb = load_workbook(myfile,read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
            name = row[0].value
            amt_bal = Money(Decimal(row[2].value),'INR')
            gold_bal = Money(Decimal(row[3].value),'USD')
            cb = Balance([amt_bal,gold_bal])
            try:
                acc = Account.objects.get(contact__name = name)

            except Account.DoesNotExist:
                logger.warning("Account Doesnot Exist")
            AccountStatement.objects.create(AccountNo = acc,ClosingBalance = cb.monies(),
                TotalCredit = Balance().monies(),TotalDebit = Balance().monies())            
        
    return render(request, 'sales/simpleupload.html')
