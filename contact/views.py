from openpyxl.reader.excel import load_workbook
from contact.utils import eliminate_dups
from django.views.generic import DetailView, UpdateView, CreateView,DeleteView
from django_tables2.views import SingleTableMixin
from django_tables2.export.views import ExportMixin
from .tables import CustomerTable
from django_filters.views import FilterView
from .filters import CustomerFilter
from .models import Customer
from .forms import CustomerForm
from django.urls import reverse_lazy
from django.shortcuts import render,redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from sales.models import Month
from django.db.models import  Sum,Q,Count
from datetime import datetime, timedelta
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.messages.views import SuccessMessageMixin
import logging
import tablib
from .admin import CustomerResource
from import_export import resources
from decimal import Decimal
logger = logging.getLogger(__name__)


def simple_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']

        ds = tablib.Dataset()
        ds.xlsx = myfile

        logger.warning(f"initial {type(ds)} height : {ds.height}")
        uniq = eliminate_dups(ds)
        logger.warning(f"unique {type(uniq) }height :{uniq.height}")
        
        customer_resource = resources.modelresource_factory(model = Customer)()
        result = customer_resource.import_data(uniq,dry_run=True,raise_errors=True)
        if not result.has_errors():
            logger.warning("no errors in import")
            customer_resource.import_data(uniq,dry_run = False,raise_errors=True)
        else:
            logger.warning("Error importing")

    return render(request,'contact/simpleupload.html')

def import_txns(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']

        wb = load_workbook(myfile,read_only=True)
        ws = wb.active

        
    return render(request,'contact/simpleupload.html')
    
def import_purctxns(request):
    
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        wb = load_workbook(myfile,read_only = True)
        ws = wb.active
        from purchase.models import Invoice as pinv
         
        ds = tablib.Dataset(headers = ['supplier','created','created_by','balancetype','rate',
                            'gross_wt','net_wt','balance','term'])
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
            
            contact = Customer.objects.get_or_create(name = row[3].value,
                        defaults={'phonenumber': 911,
                        'type':'Su',
                        'relatedas':'S/O'},)[0]
            amt = int(row[4].value)
            gold = Decimal(row[7].value)
            g_wt = Decimal(row[5].value)
            n_wt = Decimal(row[6].value)
            if amt and not gold:
                rate = round(amt/n_wt,0)
                ds.append([contact.id, row[2].value, request.user.id, 'INR',rate,g_wt,n_wt, amt, 1])
            elif gold and not amt:
                ds.append([contact.id, row[2].value, request.user.id, 'USD',0,g_wt,n_wt, gold, 1])
            elif amt and gold:
                rate = round(amt/n_wt, 0)
                ds.append([contact.id,row[2].value, request.user.id,
                          'INR', 0, 0, 0, amt, 1])
                ds.append([contact.id, row[2].value,
                          request.user.id, 'USD',0,g_wt,n_wt, gold, 1])

        purc_resource = resources.modelresource_factory(model = pinv)()
        result = purc_resource.import_data(
            ds, dry_run=True, raise_errors=True)
        if not result.has_errors():
            logger.warning('importing puchase')
            purc_resource.import_data(ds, dry_run=False, raise_errors=True)
        else:
            logger.error("Error importing purchases") 
      
    return render(request, 'contact/simpleupload.html')

def import_salestxns(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        wb = load_workbook(myfile,read_only = True)
        ws = wb.active
        from sales.models import Invoice as sinv
        ds = tablib.Dataset(headers = ['customer','created','created_by','balancetype',
                                'rate','gross_wt','net_wt','balance','term'])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            contact = Customer.objects.get_or_create(name=row[2].value,
                        defaults={'phonenumber': 911,'type': 'Wh',
                                    'relatedas': 'S/O'},)[0]
            amt = int(row[3].value)
            gold = Decimal(row[7].value)
            g_wt = row[4].value
            n_wt = row[6].value
            if amt and not gold:
                rate = round(amt/n_wt if n_wt else 1,0)
                ds.append([contact.id,row[0].value,request.user.id,'INR',rate,g_wt,n_wt,amt,1])
            elif gold and not amt:
                ds.append([contact.id,row[0].value,request.user.id,'USD',0,g_wt,n_wt,gold,1])
            elif gold and amt:
                ds.append([contact.id,row[0].value,request.user.id,'USD',0,g_wt,n_wt,gold,1])
                ds.append([contact.id, row[0].value, request.user.id,
                          'INR', 0, 0, 0, amt, 1])

        sales_resource = resources.modelresource_factory(model=sinv)()
        result = sales_resource.import_data(
            ds, dry_run=True, raise_errors=True)
        if not result.has_errors():
            logger.warning('importing sales')
            sales_resource.import_data(ds, dry_run=False, raise_errors=True)
        else:
            logger.error("Error importing sales")
            
    return render(request, 'contact/simpleupload.html')

def import_receipttxns(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        wb = load_workbook(myfile,read_only = True)
        ws = wb.active
        from sales.models import Receipt
        ds = tablib.Dataset(headers = ['customer','created','created_by','type','total','rate'])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            contact = Customer.objects.get_or_create(name=row[1].value)[0]
            amt = int(row[6].value if row[6].value else 0)
            gold = Decimal(row[2].value if row[2].value else 0)
            rate = int(row[4].value if row[4].value else 0)
            if amt and not gold:
                baltype = 'INR'
                ds.append([contact.id,row[0].value,request.user.id,baltype,amt,rate])
            else:
                baltype = 'USD'
                ds.append([contact.id, row[0].value,request.user.id, baltype, gold,rate])
        rcpt_resource = resources.modelresource_factory(model=Receipt)()
        result = rcpt_resource.import_data(
            ds, dry_run=True, raise_errors=True)
        if not result.has_errors():
            logger.warning('importing receipts')
            rcpt_resource.import_data(ds, dry_run=False, raise_errors=True)
        else:
            logger.error("Error importing receipts")
    return render(request, 'contact/simpleupload.html')

def import_paymenttxns(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        wb = load_workbook(myfile,read_only=True)
        ws = wb.active
        from purchase.models import Payment
        ds = tablib.Dataset(
            headers=['supplier', 'created', 'created_by', 'type', 'total',])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            contact = Customer.objects.get_or_create(name=row[1].value)[0]
            amt = int(row[2].value if row[2].value else 0)
            gold = round(Decimal(row[5].value if row[5].value else 0),3)
            
            if amt and not gold:
                ds.append([contact.id, row[0].value,
                          request.user.id, 'INR', amt])
            else:
                ds.append([contact.id, row[0].value,
                          request.user.id, 'USD', gold])
        payt_resource = resources.modelresource_factory(model=Payment)()
        result = payt_resource.import_data(
            ds, dry_run=True, raise_errors=True)
        if not result.has_errors():
            logger.warning('importing payments')
            payt_resource.import_data(ds, dry_run=False, raise_errors=True)
        else:
            logger.error("Error importing Payments")

    return render(request, 'contact/simpleupload.html')

@login_required
def home(request):
    data = dict()
    c=Customer.objects
    data['retailcount']=c.filter(type='Re').count()

    data['wol']=c.filter(type='Re',loan=None).count()
    data['wl']=data['retailcount']-data['wol']
    data['customercount']=c.all().count()

    data['whcount']=c.filter(type="Wh").count()
    data['withph'] = round((c.exclude(phonenumber = '911').count()/c.count())*100,2)
    data['religionwise']= c.values('religion').annotate(Count('religion')).order_by('religion')

    return render(request,'contact/home.html',context={'data':data},)

class CustomerListView(LoginRequiredMixin,ExportMixin,SingleTableMixin,FilterView):
    table_class = CustomerTable
    model = Customer
    template_name = 'contact/customer_list.html'
    filterset_class = CustomerFilter
    paginate_by = 25

class CustomerCreateView(LoginRequiredMixin,SuccessMessageMixin,CreateView):
    model = Customer
    form_class = CustomerForm
    success_url=reverse_lazy('contact_customer_list')
    success_message = "%(calculated_field)s was created successfully"

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            cleaned_data,
            calculated_field=self.object,
        )
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

def reallot_receipts(request,pk):
    customer = Customer.objects.get(pk = pk)
    customer.reallot_receipts()
    return redirect(customer.get_absolute_url())

def reallot_payments(request, pk):
    customer = Customer.objects.get(pk=pk)
    customer.reallot_payments()
    return redirect(customer.get_absolute_url())
    
class CustomerDetailView(LoginRequiredMixin,DetailView):
    model = Customer
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        data=self.object.sales.all()
        inv =data.exclude(status="Paid")
        how_many_days = 30
        context['current'] = inv.filter(created__gte = datetime.now()-timedelta(days=how_many_days)).aggregate(tc = Sum('balance',filter = Q(balancetype='Cash')),tm = Sum('balance',filter = Q(balancetype = 'Gold')))
        context['past'] = inv.filter(created__lte = datetime.now()-timedelta(days=how_many_days)).aggregate(tc = Sum('balance',filter = Q(balancetype='Cash')),tm = Sum('balance',filter = Q(balancetype = 'Gold')))
        context['monthwise'] = inv.annotate(month = Month('created')).values('month').order_by('month').annotate(tc = Sum('balance',filter = Q(balancetype='Cash')),tm = Sum('balance',filter = Q(balancetype = 'Gold'))).values('month','tm','tc')
        context['monthwiserev'] = data.annotate(month = Month('created')).values('month').order_by('month').annotate(tc = Sum('balance',filter = Q(balancetype='Cash')),tm = Sum('balance',filter = Q(balancetype = 'Gold'))).values('month','tm','tc')
        return context

class CustomerUpdateView(LoginRequiredMixin,UpdateView):
    model = Customer
    form_class = CustomerForm

class CustomerDelete(LoginRequiredMixin,DeleteView):
    model=Customer
    success_url = reverse_lazy('contact_customer_list')
